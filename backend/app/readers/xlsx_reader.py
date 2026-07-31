from pathlib import Path

from openpyxl import load_workbook

from app.readers.base import DocumentExtractionError, SourceSection


class XlsxDocumentReader:
    extensions = frozenset({".xlsx"})

    def extract(self, path: Path) -> list[SourceSection]:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
            sections: list[SourceSection] = []
            sequence = 1
            for worksheet in workbook.worksheets:
                for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                    values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                    if values:
                        sections.append(
                            SourceSection(
                                sequence=sequence,
                                text=" | ".join(values),
                                source_locator=f'Excel sheet "{worksheet.title}" rows {row_number}-{row_number}',
                                metadata={"atomic": "true"},
                            )
                        )
                        sequence += 1
            workbook.close()
        except Exception as error:
            raise DocumentExtractionError("Unable to extract displayed values from this Excel workbook.") from error

        if not sections:
            raise DocumentExtractionError("This Excel workbook has no extractable displayed values.")
        return sections
